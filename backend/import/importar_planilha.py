#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Importador da planilha de férias do RH → dados para o backend (Go Férias!).

Lê a planilha no MESMO formato do modelo (OFF/MODELO PLANILHA FÉRIAS.xlsx):
blocos por funcionário, cada linha um período aquisitivo, férias contábeis nas
colunas H/I/J e folgas (texto livre) na coluna K.

O que ele faz:
  • extrai de forma CONFIÁVEL a parte estruturada (pessoa, admissão, períodos,
    situação, férias contábeis);
  • faz um parse BEST-EFFORT das folgas (texto livre) e MARCA para revisão
    tudo que for ambíguo (venda pela MP, banco de horas, sem data, etc.);
  • gera 3 saídas: dados.json, inserts.sql e REVISAR.txt.

Uso:
  python3 importar_planilha.py "../../OFF/MODELO PLANILHA FÉRIAS.xlsx" saida/

⚠️ As folgas em texto livre NÃO têm parse 100% confiável — sempre revise o
REVISAR.txt antes de considerar a migração pronta.
"""
import sys, os, re, json, unicodedata

try:
    import openpyxl
except ImportError:
    sys.exit("Instale o openpyxl:  pip3 install openpyxl")

# ---- situação da planilha → enum do app ----
SITUACAO_MAP = {
    'FÉRIAS NÃO VENCIDAS': 'acumulando',
    'FERIAS NAO VENCIDAS': 'acumulando',
    'PAGO': 'pago',
    'PAGTO PROGRAMADO': 'programado',
    'PAGAMENTO PROGRAMADO': 'programado',
}

def norm(s):
    s = '' if s is None else str(s)
    s = unicodedata.normalize('NFD', s).encode('ascii', 'ignore').decode()
    return s.strip().upper()

# conectores que ficam minúsculos no meio do nome (não no começo)
_CONECTORES = {'de', 'da', 'do', 'das', 'dos', 'e', 'di', 'du', 'del', 'la'}

def titulo(nome):
    """'DAVI ALVES (BAHAMAS)' -> 'Davi Alves (Bahamas)'. Conectores minúsculos.
    Capitaliza a 1ª letra de cada palavra (mesmo após '(' ou '-') e abaixa o resto."""
    def cap(w):
        for i, ch in enumerate(w):
            if ch.isalpha():
                return w[:i] + ch.upper() + w[i + 1:].lower()
        return w.lower()
    def cap_hifen(w):                       # trata "ANA-MARIA" -> "Ana-Maria"
        return '-'.join(cap(x) for x in w.split('-'))
    partes = str(nome).strip().split()
    out = []
    for i, w in enumerate(partes):
        base = norm(w).lower()
        if i > 0 and base in _CONECTORES:
            out.append(base)
        else:
            out.append(cap_hifen(w))
    return ' '.join(out)

def iso_data(dia, mes, ano):
    if ano < 100:                      # ano com 2 dígitos → 20YY
        ano += 2000
    # valida: datas absurdas (mês>12, dia>31) são texto ambíguo → devolve None
    # para o segmento cair no REVISAR em vez de derrubar o parser.
    if not (1 <= mes <= 12) or not (1 <= dia <= 31):
        return None
    return f"{ano:04d}-{mes:02d}-{dia:02d}"

def parse_token_data(tok, ref_mes=None, ref_ano=None):
    """'24/06', '27/06/2025', '6/01/2022' → 'YYYY-MM-DD'. Empresta mês/ano do contexto."""
    tok = tok.strip()
    m = re.match(r'^(\d{1,2})(?:/(\d{1,2}))?(?:/(\d{2,4}))?$', tok)
    if not m:
        return None
    dia = int(m.group(1))
    mes = int(m.group(2)) if m.group(2) else ref_mes
    ano = int(m.group(3)) if m.group(3) else ref_ano
    if not mes or not ano:
        return None
    try:
        return iso_data(dia, mes, ano)
    except Exception:
        return None

def parse_intervalo_ferias(txt):
    """
    Colunas H/I/J (férias contábeis). Ex.:
      '30/07/2025 à 28/08/2025 = 30 dias'
      '02/09 a 01/10/2024 (30 dias)'
      '01/11/2023 A 30/11/2023'
      '02 a 09/09/2024 (8 dias)'
    Retorna dict {inicio, fim, dias} ou None (e a razão vai pro chamador).
    """
    if not txt:
        return None
    t = str(txt).strip()   # a célula pode vir como datetime; vira string (data solta não é intervalo → None)
    if norm(t) in ('X', ''):
        return None
    # precisa de duas datas separadas por a/à/A
    m = re.search(r'(\d{1,2}(?:/\d{1,2})?(?:/\d{2,4})?)\s*[aàAÀ]\s*(\d{1,2}(?:/\d{1,2})?(?:/\d{2,4})?)', t)
    if not m:
        return None            # ex.: "Ver data de agendamento" → não é intervalo
    fim = parse_token_data(m.group(2))
    if not fim:
        return None
    fa = int(fim[:4]); fm = int(fim[5:7])
    ini = parse_token_data(m.group(1), ref_mes=fm, ref_ano=fa)
    if not ini:
        return None
    # dias: explícito "(N dias)" / "= N dias", senão calcula pelo intervalo (inclusive)
    dm = re.search(r'(\d{1,3})\s*dias', t, re.I) or re.search(r'\((\d{1,3})\)', t)
    if dm:
        dias = int(dm.group(1))
    else:
        from datetime import date
        d0 = date(int(ini[:4]), int(ini[5:7]), int(ini[8:10]))
        d1 = date(fa, fm, int(fim[8:10]))
        dias = (d1 - d0).days + 1
    return {'inicio': ini, 'fim': fim, 'dias': dias}

# frases que indicam que NÃO é data (precisam de decisão humana)
PADROES_REVISAR = re.compile(r'\b(MP|VEND|BANCO|LICEN|MATERNIDADE|DESCONTAR|AGENDAMENTO)\b', re.I)

def parse_folgas(txt):
    """
    Coluna K (folgas em texto livre). Best-effort:
      • quebra por '+' e '-' em segmentos;
      • de cada segmento tenta extrair uma data/intervalo e uma contagem;
      • segmentos sem data OU com frase de negócio vão para 'revisar'.
    Retorna (lista_folgas, lista_revisar).
    """
    folgas, revisar = [], []
    if not txt or not str(txt).strip():
        return folgas, revisar
    t = str(txt)
    # remove um "= N dias/DIAS" final (é só o somatório da linha)
    t = re.sub(r'=\s*\d{1,3}\s*(dias|fds)?.*$', '', t, flags=re.I)
    # protege '+' e '-' DENTRO de parênteses (ex.: "(14 + 1)") antes de quebrar
    t = _protege_parens(t)
    segmentos = re.split(r'[+\-]', t)
    for seg in segmentos:
        seg = seg.replace('\x00', '+').replace('\x01', '-').strip()   # restaura
        if not seg:
            continue
        if PADROES_REVISAR.search(seg):
            revisar.append(seg); continue
        # intervalo "DD.. a/à DD/MM/AAAA"
        mi = re.search(r'(\d{1,2}(?:/\d{1,2})?)\s*[aàAÀ]\s*(\d{1,2}/\d{1,2}(?:/\d{2,4})?)', seg)
        # datas soltas "DD/MM/AAAA" ou "DD,DD/MM" ou "DD e DD/MM/AAAA"
        datas = re.findall(r'\d{1,2}/\d{1,2}(?:/\d{2,4})?', seg)
        cont = re.search(r'\((\s*\d{1,3}[^\)]*)\)', seg)   # "(6)" / "(14 + 1)" / "(7+3 )"
        obs = re.sub(r'\s+', ' ', seg).strip()
        if mi:
            fim = parse_token_data(mi.group(2))
            if fim:
                fa, fm = int(fim[:4]), int(fim[5:7])
                ini = parse_token_data(mi.group(1), ref_mes=fm, ref_ano=fa)
                dias = _soma_cont(cont) or _dias_intervalo(ini, fim)
                if ini:
                    motivo = _valida_folga(ini, fim, dias)
                    if motivo:
                        revisar.append(f'{obs}  [!{motivo}]')
                    else:
                        folgas.append({'inicio': ini, 'fim': fim, 'dias': dias, 'obs': _limpa_obs(cont, seg)})
                    continue
        if datas:
            # usa a última data como referência de mês/ano
            ref = parse_token_data(datas[-1])
            if ref:
                ra, rm = int(ref[:4]), int(ref[5:7])
                todas = [parse_token_data(d, ref_mes=rm, ref_ano=ra) for d in datas]
                todas = [d for d in todas if d]
                if todas:
                    dias = _soma_cont(cont) or len(todas)
                    ini2, fim2 = min(todas), max(todas)
                    motivo = _valida_folga(ini2, fim2, dias)
                    if motivo:
                        revisar.append(f'{obs}  [!{motivo}]')
                    else:
                        folgas.append({'inicio': ini2, 'fim': fim2, 'dias': dias, 'obs': _limpa_obs(cont, seg)})
                    continue
        # não conseguiu extrair data → revisar
        revisar.append(obs)
    return folgas, revisar

def _protege_parens(t):
    """Troca '+'/'-' dentro de parênteses por sentinelas, p/ não quebrar contagens."""
    out, depth = [], 0
    for ch in t:
        if ch == '(':
            depth += 1
        elif ch == ')':
            depth = max(0, depth - 1)
        if depth > 0 and ch == '+':
            out.append('\x00')
        elif depth > 0 and ch == '-':
            out.append('\x01')
        else:
            out.append(ch)
    return ''.join(out)

def _soma_cont(cont):
    if not cont:
        return None
    nums = re.findall(r'\d{1,3}', cont.group(1))
    return sum(int(n) for n in nums) if nums else None

def _dias_intervalo(ini, fim):
    if not ini or not fim:
        return 1
    from datetime import date
    d0 = date(int(ini[:4]), int(ini[5:7]), int(ini[8:10]))
    d1 = date(int(fim[:4]), int(fim[5:7]), int(fim[8:10]))
    return max((d1 - d0).days + 1, 1)

def _limpa_obs(cont, seg):
    # observação curta preservando a contagem original entre parênteses
    return (cont.group(1).strip() if cont else re.sub(r'\s+', ' ', seg).strip())[:60]

def _valida_folga(ini, fim, dias):
    """Retorna None se a folga é coerente, senão o motivo p/ mandar ao REVISAR."""
    from datetime import date
    try:
        d0 = date(int(ini[:4]), int(ini[5:7]), int(ini[8:10]))
        d1 = date(int(fim[:4]), int(fim[5:7]), int(fim[8:10]))
    except Exception:
        return 'data inválida'
    if d1 < d0:
        return 'fim antes do início'
    span = (d1 - d0).days + 1
    if dias and dias > span:              # não dá pra tirar mais dias que o intervalo comporta
        return f'{dias} dias num intervalo de {span}'
    return None

def revisar_incoerencias(p):
    """Checagens CRUZADAS por pessoa: move p/ 'revisar' o que não dá pra confiar,
    em vez de importar lixo (período degenerado/duplicado). Férias contábeis e
    folga são independentes — folga coincidente com férias NÃO é removida.

    NÃO comparamos folga com a admissão: no grupo há troca de CNPJ (readmissão),
    então períodos/folgas legítimos podem ser anteriores à admissão atual."""
    from datetime import date
    def d(iso):
        try:
            return date(int(iso[:4]), int(iso[5:7]), int(iso[8:10]))
        except Exception:
            return None
    vistos = set()
    inicios = set()
    for per in p['periodos']:
        pini, pfim = d(per['inicio']), d(per['fim'])
        chave = (per['inicio'], per['fim'])
        if pini and pfim and pfim <= pini:                       # 0/1 dia → lixo de readmissão
            per['revisar'].append({'linha': 0, 'texto': f"período de 0/1 dia {per['inicio']}..{per['fim']} — NÃO importado"})
            per['pular'] = True
        elif chave in vistos:                                    # duplicado exato → lixo
            per['revisar'].append({'linha': 0, 'texto': f"período duplicado {per['inicio']}..{per['fim']} — NÃO importado"})
            per['pular'] = True
        elif per['inicio'] in inicios:                           # mesmo início, fim diferente → conferir
            per['revisar'].append({'linha': 0, 'texto': f"período com mesmo início de outro ({per['inicio']}) — conferir no app"})
        vistos.add(chave)
        inicios.add(per['inicio'])
        # Férias contábeis e folga são INDEPENDENTES (na férias contábil a pessoa trabalha
        # normal, sem bater ponto). Uma folga coincidente com férias NÃO é dupla contagem —
        # mantemos ambas.

def parse_planilha(caminho):
    wb = openpyxl.load_workbook(caminho, data_only=True)
    ws = wb.active
    pessoas = []
    atual = None
    for r in range(1, ws.max_row + 1):
        A = ws.cell(r, 1).value
        if norm(A) == 'FUNCIONARIO':          # cabeçalho (repetido por bloco)
            continue
        D = ws.cell(r, 4).value               # admissão presente em toda linha do bloco
        if A and str(A).strip():              # início de um novo funcionário
            atual = {
                'nome': titulo(A),
                'empresa': (str(ws.cell(r, 2).value).strip() if ws.cell(r, 2).value else ''),
                'funcao': (titulo(str(ws.cell(r, 3).value).strip()) if ws.cell(r, 3).value else ''),
                'admissao': (D.date().isoformat() if hasattr(D, 'date') else str(D)[:10]) if D else None,
                'periodos': [],
            }
            pessoas.append(atual)
        if atual is None:
            continue
        E, F, G = ws.cell(r, 5).value, ws.cell(r, 6).value, ws.cell(r, 7).value
        if not E:                             # linha sem período aquisitivo → pula
            continue
        per = {
            'inicio': E.date().isoformat() if hasattr(E, 'date') else str(E)[:10],
            'fim':    F.date().isoformat() if hasattr(F, 'date') else str(F)[:10],
            'situacao': SITUACAO_MAP.get(norm(G), None),
            'situacao_original': (str(G).strip() if G else ''),
            'ferias': [],
            'folgas': [],
            'revisar': [],
        }
        for col in (8, 9, 10):                # H/I/J → férias contábeis
            fx = parse_intervalo_ferias(ws.cell(r, col).value)
            if fx:
                per['ferias'].append(fx)
        K = ws.cell(r, 11).value              # folgas
        fol, rev = parse_folgas(K)
        per['folgas'] = fol
        if rev:
            per['revisar'] = [{'linha': r, 'texto': x} for x in rev]
        atual['periodos'].append(per)
    for p in pessoas:
        revisar_incoerencias(p)
    return pessoas

def gerar_sql(pessoas):
    """Gera INSERTs para o schema `api` (uuid geradas pelo banco)."""
    def q(s):
        return "'" + str(s).replace("'", "''") + "'" if s is not None else 'null'
    out = ["-- Gerado por importar_planilha.py — revise o REVISAR.txt antes de usar.",
           "begin;"]
    for p in pessoas:
        out.append(f"\nwith c as (insert into api.colaboradores(nome,funcao,departamento,unidade,regime,admissao) "
                   f"values ({q(p['nome'])},{q(p['funcao'])},{q(p['empresa'])},'',{q('CLT')},{q(p['admissao'])}) returning id)")
        # cada período com suas férias/folgas
        blocos = []
        for i, per in enumerate(p['periodos']):
            pv = f"p{i}"
            blocos.append(f"{pv} as (insert into api.periodos_aquisitivos(colaborador_id,inicio,fim,situacao) "
                          f"select id,{q(per['inicio'])},{q(per['fim'])},{q(per['situacao'] or 'acumulando')} from c returning id)")
        if blocos:
            out.append(", " + ", ".join(blocos))
        # inserts de ferias/folgas referenciando pX
        linhas = []
        for i, per in enumerate(p['periodos']):
            pv = f"p{i}"
            for f in per['ferias']:
                linhas.append(f"insert into api.ferias_oficiais(periodo_id,inicio,fim,dias) select id,{q(f['inicio'])},{q(f['fim'])},{f['dias']} from {pv};")
            for f in per['folgas']:
                linhas.append(f"insert into api.folgas(periodo_id,inicio,fim,dias,obs) select id,{q(f['inicio'])},{q(f['fim'])},{f['dias']},{q(f.get('obs',''))} from {pv};")
        out.append(";\n" + "\n".join(linhas) if linhas else ";")
    out.append("\ncommit;")
    return "\n".join(out)

def main():
    if len(sys.argv) < 2:
        sys.exit(__doc__)
    caminho = sys.argv[1]
    saida = sys.argv[2] if len(sys.argv) > 2 else 'saida'
    os.makedirs(saida, exist_ok=True)
    pessoas = parse_planilha(caminho)

    with open(os.path.join(saida, 'dados.json'), 'w', encoding='utf-8') as f:
        json.dump(pessoas, f, ensure_ascii=False, indent=2)
    with open(os.path.join(saida, 'inserts.sql'), 'w', encoding='utf-8') as f:
        f.write(gerar_sql(pessoas))

    # relatório de revisão
    total_rev = 0
    with open(os.path.join(saida, 'REVISAR.txt'), 'w', encoding='utf-8') as f:
        f.write("SEGMENTOS DE FOLGA QUE PRECISAM DE REVISÃO HUMANA\n")
        f.write("(sem data reconhecível ou com regra de negócio — MP, banco de horas, etc.)\n\n")
        for p in pessoas:
            revs = [(per, x) for per in p['periodos'] for x in per['revisar']]
            if revs:
                f.write(f"== {p['nome']} ==\n")
                for per, x in revs:
                    total_rev += 1
                    f.write(f"  período {per['inicio']}..{per['fim']}: {x['texto']}\n")
                f.write("\n")

    # resumo no terminal
    print(f"Funcionários: {len(pessoas)}")
    for p in pessoas:
        nper = len(p['periodos'])
        nfer = sum(len(x['ferias']) for x in p['periodos'])
        nfol = sum(len(x['folgas']) for x in p['periodos'])
        nrev = sum(len(x['revisar']) for x in p['periodos'])
        print(f"  {p['nome']:<32} períodos:{nper}  férias:{nfer}  folgas:{nfol}  revisar:{nrev}")
    print(f"\nTotal a revisar: {total_rev}")
    print(f"Saídas em: {saida}/  (dados.json, inserts.sql, REVISAR.txt)")

if __name__ == '__main__':
    main()
