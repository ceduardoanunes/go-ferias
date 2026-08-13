#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Importador do FORM 18 (planilha REAL do RH, multi-aba) -> dados.json p/ o backend.

Diferente do importar_planilha.py (feito p/ o MODELO de 2 pessoas numa aba só), este:
  - percorre TODAS as abas de setor ATIVO (uma por setor);
  - ignora as abas de desligados e as abas-meta (LEGENDA, SALDOS, etc.);
  - limpa lixo real do arquivo:
      * linhas de TÍTULO da aba ("CONTROLE DE FÉRIAS CONTÁBEIS - CRIAÇÃO");
      * linhas de NOTA/instrução (texto longo na coluna do nome);
      * placeholders genéricos (ex.: "ESTÁGIÁRIO");
      * períodos cujo início não é uma data (artefato de cabeçalho repetido);
      * datas em TEXTO ("DD/MM/AAAA"), além das já-datetime;
  - separa quem NÃO tem admissão válida (a API exige admissao) em PENDENTES.txt
    — esses NÃO são enviados; o RH lança à mão depois;
  - reaproveita o parse de FÉRIAS CONTÁBEIS (H/I/J) e de FOLGAS (K, texto livre)
    do importar_planilha.py, com o REVISAR.txt para os casos de negócio.

Uso:
  python3 importar_form18.py "../../FORM 18 - ... .xlsx" saida
  python3 importar_form18.py "../../FORM 18 - ... .xlsx" saida --aba "VÍDEO"   # só um setor
"""
import sys, os, re, json
import openpyxl
from importar_planilha import (
    norm, titulo, SITUACAO_MAP, parse_intervalo_ferias, parse_folgas,
    revisar_incoerencias,
)

# Abas que NÃO são de colaboradores ativos.
META = {
    'SALDOS DE FÉRIAS', 'LEGENDA', 'Planilha2', 'Ciclos fechados',
    'AGENDAMENTOS SET A OUT 2024', 'Plan3', 'BAYER, NEXA E NISSAN',
}
DESLIGADOS = {
    'EX FUNCIONÁRIOS', 'DESLIGADOS RECENTES', 'DESLIGADOS 2026', 'DESLIGADOS ANTIGOS',
}

# Nome da ABA (normalizado) -> setor do app. Definido pelo RH (apurado no front).
# REDAÇÃO-REVISÃO é dividida por FUNÇÃO (ver setor_de).
SETOR_MAP = {
    'ATEND. SOCIAL MIDIA': 'Digital',
    'ATENDIMENTO OFF': 'Atendimento OFF',
    'ATENDIMENTO CORPORATIVO': 'Corporativo',
    'SUPERVISAO': 'Coordenação',
    'CRIACAO': 'Criação',
    'VIDEO': 'Vídeo',
    'INBOUND': 'Inbound',
    'ADM': 'Administrativo',
    'ESTAGIARIOS': 'Estagiários',
}

def setor_de(aba, funcao):
    """Setor do app a partir da aba (e da função, no caso Redação-Revisão)."""
    na = norm(aba)
    if 'REDA' in na and 'REVIS' in na:                 # aba REDAÇÃO-REVISÃO
        return 'Revisão' if 'REVIS' in norm(funcao) else 'Redação'
    return SETOR_MAP.get(na)                            # None se aba não mapeada

def to_iso(v):
    """datetime OU texto ('DD/MM/AAAA', 'DD/MM/AA', 'AAAA-MM-DD') -> 'AAAA-MM-DD' ou None."""
    if v is None:
        return None
    if hasattr(v, 'date'):
        return v.date().isoformat()
    s = str(v).strip()
    if not s:
        return None
    m = re.match(r'^(\d{4})-(\d{2})-(\d{2})', s)      # já ISO (com ou sem hora)
    if m:
        y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if 1 <= mo <= 12 and 1 <= d <= 31:
            return f'{y:04d}-{mo:02d}-{d:02d}'
    m = re.match(r'^(\d{1,2})/(\d{1,2})/(\d{2,4})$', s)  # DD/MM/AAAA
    if m:
        d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if y < 100:
            y += 2000
        if 1 <= mo <= 12 and 1 <= d <= 31:
            return f'{y:04d}-{mo:02d}-{d:02d}'
    return None

def eh_titulo_ou_nota(nome):
    """True p/ linhas que NÃO são pessoa: título da aba, nota longa, placeholder."""
    n = norm(nome)
    if not n:
        return True
    if n.startswith('CONTROLE') or 'CONTABEIS' in n or 'CONTABIL' in n:
        return True
    if n in ('ESTAGIARIO', 'ESTAGIARIA', 'FUNCIONARIO', 'X', '-'):
        return True
    if len(str(nome).strip()) > 60:      # instrução/nota (ex.: bloco da BAYER)
        return True
    return False

def parse_aba(ws, setor):
    pessoas = []
    atual = None
    for r in range(1, ws.max_row + 1):
        A = ws.cell(r, 1).value
        if norm(A) == 'FUNCIONARIO':
            continue
        if A and str(A).strip():                    # possível início de pessoa
            if eh_titulo_ou_nota(A):
                atual = None                        # descarta bloco de título/nota
                continue
            D = ws.cell(r, 4).value
            funcao = (str(ws.cell(r, 3).value).strip() if ws.cell(r, 3).value else '')
            atual = {
                'nome': titulo(A),
                'empresa': (str(ws.cell(r, 2).value).strip() if ws.cell(r, 2).value else ''),
                'funcao': titulo(funcao),
                'admissao': to_iso(D),
                'setor_aba': setor,
                'setor': setor_de(setor, funcao),        # setor do app (departamento)
                'periodos': [],
            }
            pessoas.append(atual)
        if atual is None:
            continue
        E = ws.cell(r, 5).value
        ini = to_iso(E)
        if not ini:                                 # período sem início-data válido -> pula
            continue
        F, G = ws.cell(r, 6).value, ws.cell(r, 7).value
        per = {
            'inicio': ini,
            'fim': to_iso(F) or ini,
            'situacao': SITUACAO_MAP.get(norm(G), None),
            'situacao_original': (str(G).strip() if G else ''),
            'ferias': [], 'folgas': [], 'revisar': [],
        }
        for col in (8, 9, 10):
            fx = parse_intervalo_ferias(ws.cell(r, col).value)
            if fx:
                per['ferias'].append(fx)
        fol, rev = parse_folgas(ws.cell(r, 11).value)
        per['folgas'] = fol
        if rev:
            per['revisar'] = [{'linha': r, 'texto': x} for x in rev]
        atual['periodos'].append(per)
    return pessoas

def main():
    # --aba "NOME": processa só um setor por vez (ignora acento/caixa).
    args = sys.argv[1:]
    aba_filtro = None
    if '--aba' in args:
        i = args.index('--aba')
        aba_filtro = args[i + 1] if i + 1 < len(args) else None
        del args[i:i + 2]
    if not args:
        sys.exit(__doc__)
    caminho = args[0]
    saida = args[1] if len(args) > 1 else 'saida'
    os.makedirs(saida, exist_ok=True)

    wb = openpyxl.load_workbook(caminho, data_only=True)
    todas, pendentes = [], []
    for name in wb.sheetnames:
        if name in META or name in DESLIGADOS:
            continue
        if aba_filtro and norm(name) != norm(aba_filtro):
            continue
        for p in parse_aba(wb[name], name):
            # descarta "pessoa" totalmente vazia (nenhum período e sem admissão)
            if not p['admissao'] and not p['periodos']:
                continue
            if not p['admissao']:
                pendentes.append(p)               # tem dados mas falta admissão (API exige)
            else:
                todas.append(p)

    for p in todas:                       # checagens cruzadas: tira lixo, joga suspeito p/ REVISAR
        revisar_incoerencias(p)

    with open(os.path.join(saida, 'dados.json'), 'w', encoding='utf-8') as f:
        json.dump(todas, f, ensure_ascii=False, indent=2)

    # PENDENTES: pessoas com dados mas sem admissão válida -> lançar à mão
    with open(os.path.join(saida, 'PENDENTES.txt'), 'w', encoding='utf-8') as f:
        f.write("PESSOAS NÃO ENVIADAS — falta ADMISSÃO válida (a API exige).\n")
        f.write("Lance à mão no app (ou corrija a planilha e reimporte).\n\n")
        for p in pendentes:
            f.write(f"== {p['nome']}  [{p['setor_aba']}]  função: {p['funcao']}  períodos: {len(p['periodos'])}\n")

    # REVISAR: folgas em texto livre com regra de negócio
    total_rev = 0
    with open(os.path.join(saida, 'REVISAR.txt'), 'w', encoding='utf-8') as f:
        f.write("SEGMENTOS DE FOLGA QUE PRECISAM DE REVISÃO HUMANA\n")
        f.write("(sem data reconhecível ou regra de negócio — MP, banco de horas, licença…)\n\n")
        for p in todas:
            revs = [(per, x) for per in p['periodos'] for x in per['revisar']]
            if revs:
                f.write(f"== {p['nome']} [{p['setor_aba']}] ==\n")
                for per, x in revs:
                    total_rev += 1
                    f.write(f"  período {per['inicio']}..{per['fim']}: {x['texto']}\n")
                f.write("\n")

    # resumo
    nper = sum(len(p['periodos']) for p in todas)
    nfer = sum(len(x['ferias']) for p in todas for x in p['periodos'])
    nfol = sum(len(x['folgas']) for p in todas for x in p['periodos'])
    print(f"Colaboradores a ENVIAR: {len(todas)}")
    print(f"  períodos: {nper}   férias contábeis: {nfer}   folgas: {nfol}")
    print(f"Pendentes (sem admissão, NÃO enviados): {len(pendentes)}")
    print(f"Folgas a revisar (à mão): {total_rev}")
    print(f"\nSaídas em: {saida}/  (dados.json, PENDENTES.txt, REVISAR.txt)")

if __name__ == '__main__':
    main()
